"""
convert.py — runs automatically on GitHub Actions
Converts IPM Excel file → data.json for the dashboard
"""
import json, glob, os
from datetime import datetime
import openpyxl

# Find the Excel file (any .xlsx in repo root)
xlsx_files = glob.glob("*.xlsx") + glob.glob("*.xlsm")
if not xlsx_files:
    print("No Excel file found in repo root!")
    exit(1)

xlsx_path = xlsx_files[0]
print(f"Reading: {xlsx_path}")

wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

# Find IPM sheet
sheet_name = next((s for s in wb.sheetnames if 'IPM' in s.upper()), wb.sheetnames[0])
ws = wb[sheet_name]
print(f"Sheet: {sheet_name}")

def parse_date(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 3:
            d, m, y = parts
            return f"{y.zfill(4)}-{m.zfill(2)}-{d.zfill(2)}"
    return s

records = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue  # skip header
    # Skip empty rows
    if not row[0] or not row[1] or not row[2]:
        continue
    try:
        records.append({
            "date":         parse_date(row[0]),
            "gh":           int(row[1]),
            "sector":       int(row[2]),
            "valve":        str(row[3]) if row[3] else "A&B",
            "first_line":   int(row[4]) if row[4] else 1,
            "end_line":     int(row[5]) if row[5] else 160,
            "pesticide":    str(row[8])  if row[8]  else "",
            "active":       str(row[9])  if row[9]  else "",
            "method":       str(row[10]) if row[10] else "Spray",
            "phi":          int(row[15]) if row[15] else 0,
            "harvest_date": parse_date(row[16]),
            "status":       str(row[17]) if row[17] else "",
        })
    except Exception as e:
        print(f"  Skipping row {i+1}: {e}")

output = {
    "updated": datetime.utcnow().strftime("%d %b %Y %H:%M UTC"),
    "source":  xlsx_path,
    "records": records
}

with open("data.json", "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"✓ data.json written — {len(records)} records")
